#!/usr/bin/env python3
"""
Gerador de planilhas Excel para relatórios do BarberFlow
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
import sys

def format_currency(value):
    """Formata valor para moeda brasileira"""
    return f"R$ {value:,.2f}".replace(',', '_').replace('.', ',').replace('_', '.')

def format_date(date_str):
    """Formata data para padrão brasileiro"""
    try:
        dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
        return dt.strftime('%d/%m/%Y %H:%M')
    except:
        return date_str

def generate_appointments_excel(data, output_path):
    """Gera Excel de relatório de agendamentos"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Agendamentos"
    
    # Estilos
    header_fill = PatternFill(start_color="6B21A8", end_color="6B21A8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    summary_fill = PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid")
    summary_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = 'RELATÓRIO DE AGENDAMENTOS'
    ws['A1'].font = Font(bold=True, size=16, color="6B21A8")
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:F1')
    
    # Período
    row = 2
    if data.get('filters', {}).get('startDate'):
        ws[f'A{row}'] = f"Período: {format_date(data['filters']['startDate'])} a {format_date(data['filters']['endDate'])}"
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
    
    row += 1  # Espaço
    
    # Resumo
    ws[f'A{row}'] = 'RESUMO'
    ws[f'A{row}'].font = summary_font
    ws[f'A{row}'].fill = summary_fill
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    summary = data.get('summary', {})
    summary_data = [
        ['Total de Agendamentos:', summary.get('total', 0)],
        ['Concluídos:', summary.get('completed', 0)],
        ['Cancelados:', summary.get('cancelled', 0)],
        ['Receita Total:', format_currency(summary.get('totalRevenue', 0))]
    ]
    
    for label, value in summary_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = summary_fill
        ws[f'B{row}'].fill = summary_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='left')
        ws[f'B{row}'].alignment = Alignment(horizontal='right')
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    row += 2  # Espaço
    
    # Cabeçalhos da tabela
    ws[f'A{row}'] = 'Data'
    ws[f'B{row}'] = 'Cliente'
    ws[f'C{row}'] = 'Telefone'
    ws[f'D{row}'] = 'Barbeiro'
    ws[f'E{row}'] = 'Serviço'
    ws[f'F{row}'] = 'Status'
    ws[f'G{row}'] = 'Valor'
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        cell = ws[f'{col}{row}']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    row += 1
    
    # Dados dos agendamentos
    for apt in data.get('appointments', []):
        ws[f'A{row}'] = format_date(apt['date'])
        ws[f'B{row}'] = apt['customer']['name']
        ws[f'C{row}'] = apt['customer']['phone']
        ws[f'D{row}'] = apt['barber']['name']
        ws[f'E{row}'] = apt['service']['name']
        ws[f'F{row}'] = apt['status']
        ws[f'G{row}'] = format_currency(float(apt['price']))
        
        # Estilo das células
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws[f'{col}{row}'].border = border
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
        
        row += 1
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    # Salvar
    wb.save(output_path)
    print(f"✅ Excel gerado: {output_path}")

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Uso: python excel-generator.py <json_file> <output_excel>")
        sys.exit(1)
    
    with open(sys.argv[1], 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    generate_appointments_excel(data, sys.argv[2])